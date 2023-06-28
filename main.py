import os
import re
from datetime import datetime
from typing import Any

import lxml.html
import openpyxl
from lxml import etree


class Fog:
    re_syllables = re.compile(r'(^|[^aeuoiy])(?!e$)[aeouiy]', re.IGNORECASE)

    @classmethod
    def count_syllables(cls, word: str):
        """Counts the number of syllables in a word."""
        syllables_matches = cls.re_syllables.findall(word)
        return len(syllables_matches)

    @classmethod
    def is_complex_word(cls, word: str):
        """Checks whether word has three or more syllables."""
        return cls.count_syllables(word) >= 3

    @classmethod
    def identify_words(cls, input_text: str):
        """Extracts all words from a given text."""
        words = re.findall(r"\b[a-zA-Z\'\-]+\b", input_text)
        return words

    @classmethod
    def count_words(cls, input_text: str):
        """Counts the number of words in a given text."""
        words = cls.identify_words(input_text)
        word_count = len(words)
        return word_count

    @classmethod
    def identify_sentences(cls, input_text: str):
        """Extracts all sentences from a given text."""
        sentences = re.findall(r"\b[A-Z](?:[^\.!?]|\.\d)*[\.!?]", input_text)
        return sentences

    @classmethod
    def count_sentences(cls, input_text: str):
        """Counts the number of sentences in input_text."""
        sentences = cls.identify_sentences(input_text)
        sentence_count = len(sentences)
        return sentence_count

    @classmethod
    def calculate_fog(cls, text: str, *args: Any, **kwds: Any) -> Any:
        """Calculates the fog index for a given text."""
        sentences = cls.identify_sentences(text)
        words = cls.identify_words(text)
        complex_words = list(filter(cls.is_complex_word, words))
        return 0.4*(float(len(words))/float(len(sentences)) + 100*float(len(complex_words))/float(len(words)))


class BaseMDNA:
    @classmethod
    def extract_mdna(cls, html_source: str):
        raise NotImplementedError()

    @classmethod
    def get_text_from_html(cls, html: str):
        """Converts HTML code to plain text"""
        doc = lxml.html.fromstring(html)

        for tag in ["tr", "th", "td", "a", "p", "div", "br", "h1", "h2", "h3", "h4", "h5"]:
            for element in doc.findall(tag):
                if element.text:
                    element.text = element.text + "\n"
                else:
                    element.text = "\n"
        return doc.text_content().strip()

    @classmethod
    def get_mdna_text(cls, html_complete_10k: str, *args: Any, **kwds: Any) -> Any:
        html_mdna_only, year = cls.extract_mdna(html_complete_10k)
        if html_mdna_only is None:
            return None, year
        mdna_text = cls.get_text_from_html(html_mdna_only)
        return mdna_text, year


class AppleMDNA(BaseMDNA):
    @classmethod
    def _extract_mdna_after_2019(cls, doc):
        for td in doc.xpath('//td[@class="text"]'):
            if td.text is None:
                continue

            if td.text.strip().startswith('Summary of Significant Accounting Policies'):
                return etree.tostring(td)
        return None

    @classmethod
    def _extract_mdna_before_2019(cls, doc):
        for div in doc.xpath('//div[@id="divSummary of Significant Accounting Policies"]/div'):
            if div.text is None:
                continue
            return etree.tostring(div)
        return None

    @classmethod
    def extract_mdna(cls, html_source: str):
        doc = lxml.html.fromstring(html_source)

        title = doc.xpath('//div[@class="ModuleFilingTitle"]/span')
        if not title or title[0].text is None:
            return
        date_str = title[0].text.split()[-1]
        report_date = datetime.strptime(date_str, '%m/%d/%Y')
        year = report_date.strftime('%Y')
        if report_date > datetime(year=2019, month=1, day=1):
            return cls._extract_mdna_after_2019(doc), year
        return cls._extract_mdna_before_2019(doc), year


class NetflixMDNA(BaseMDNA):
    @classmethod
    def extract_mdna(cls, html_source: str):
        doc = lxml.html.fromstring(html_source)
        title = doc.xpath('//div[@class="ModuleFilingTitle"]/span')
        if not title or title[0].text is None:
            return
        date_str = title[0].text.split()[-1]
        report_date = datetime.strptime(date_str, '%m/%d/%Y')
        year = report_date.strftime('%Y')
        for div in doc.xpath('//div[@class="reportContainer subcontainerNotSelected"]'):
            string = etree.tostring(div)

            if b'Organization And Summary Of Significant Accounting Policies (Policy)' in string:
                return string, year
            elif b'Organization and Summary of Significant Accounting Policies (Policy)' in string:
                return string, year
        return None, year


class FacebookMDNA(AppleMDNA):
    pass


class AmazonMDNA(BaseMDNA):
    @classmethod
    def extract_mdna(cls, html_source: str):
        doc = lxml.html.fromstring(html_source)
        title = doc.xpath('//div[@class="ModuleFilingTitle"]/span')
        if not title or title[0].text is None:
            return
        date_str = title[0].text.split()[-1]
        report_date = datetime.strptime(date_str, '%m/%d/%Y')
        year = report_date.strftime('%Y')
        for div in doc.xpath('//div[@class="reportContainer subcontainerNotSelected"]'):
            string = etree.tostring(div)
            if b'Description of Business, Accounting Policies, and Supplemental Disclosures (Policies)' in string:
                return string, year
            elif b'DESCRIPTION OF BUSINESS AND ACCOUNTING POLICIES (Policies)' in string:
                return string, year
            elif b'Description of Business and Accounting Policies (Policies)' in string:
                return string, year
        return None, year


class GoogleMDNA(BaseMDNA):
    mdan_cmp = re.compile(
        'MANAGEMENT’S DISCUSSION AND ANALYSIS OF FINANCIAL CONDITION AND RESULTS OF OPERATIONS.*?\<hr', re.DOTALL)

    @classmethod
    def extract_mdna(cls, html_source: str):

        # print(html_source)
        result = cls.mdan_cmp.search(html_source)
        if result is not None:
            result = result.group()
        return result, None


def main():
    profile = openpyxl.Workbook()
    sheet = profile['Sheet']

    title_list = ['company', 'year', 'MD&A', 'fog']
    for i, title in enumerate(title_list, 1):
        sheet.cell(1, i, title)

    row = 2
    apple_10k_files = os.listdir('./apple')
    for file in apple_10k_files:
        with open(f'apple/{file}', 'r', encoding='windows-1252') as f:
            html_complete_10k = f.read()

        mdna_text, year = AppleMDNA.get_mdna_text(html_complete_10k)
        if mdna_text is None:
            continue
        sheet.cell(row, 1, 'APPLE INC')
        sheet.cell(row, 2, year)
        sheet.cell(row, 3, mdna_text)

        fog_score = Fog.calculate_fog(mdna_text)
        sheet.cell(row, 4, fog_score)
        row += 1

    netflix_10k_files = os.listdir('./netflix')
    for file in netflix_10k_files:
        # file = "73d60966-401d-46da-9e44-53b52f450e4f.html"
        with open(f'netflix/{file}', 'r', encoding='windows-1252') as f:
            html_complete_10k = f.read()
        mdna_text, year = NetflixMDNA.get_mdna_text(html_complete_10k)
        if mdna_text is None:
            continue
        sheet.cell(row, 1, 'NETFLIX INC')
        sheet.cell(row, 2, year)
        sheet.cell(row, 3, mdna_text)

        fog_score = Fog.calculate_fog(mdna_text)
        sheet.cell(row, 4, fog_score)
        row += 1

    facebook_10k_files = os.listdir('./facebook')
    for file in facebook_10k_files:
        with open(f'facebook/{file}', encoding='windows-1252') as f:
            html_complete_10k = f.read()
            mdna_text, year = AppleMDNA.get_mdna_text(html_complete_10k)
            if mdna_text is None:
                continue
            sheet.cell(row, 1, 'FACEBOOK INC')
            sheet.cell(row, 2, year)
            sheet.cell(row, 3, mdna_text)

            fog_score = Fog.calculate_fog(mdna_text)
            sheet.cell(row, 4, fog_score)
            row += 1

    amazon_10k_files = os.listdir('./amazon')
    for file in amazon_10k_files:
        with open(f'amazon/{file}', encoding='windows-1252') as f:
            html_complete_10k = f.read()
            mdna_text, year = AmazonMDNA.get_mdna_text(html_complete_10k)
            if mdna_text is None:
                continue
            sheet.cell(row, 1, 'AMAZON COM INC')
            sheet.cell(row, 2, year)
            sheet.cell(row, 3, mdna_text)

            fog_score = Fog.calculate_fog(mdna_text)
            sheet.cell(row, 4, fog_score)
            row += 1

    google_10k_files = os.listdir('./google')
    for file in google_10k_files:
        year, *_ = file.split()
        try:
            f = open(f'google/{file}', encoding='windows-1252')
            html_complete_10k = f.read()
        except Exception as e:
            f = open(f'google/{file}', encoding='utf-8')
            html_complete_10k = f.read()
        else:
            f.close()

        mdna_text, _ = GoogleMDNA.get_mdna_text(html_complete_10k)
        if mdna_text is None:
            continue
        sheet.cell(row, 1, 'GOOGLE INC')
        sheet.cell(row, 2, year)
        sheet.cell(row, 3, mdna_text)

        fog_score = Fog.calculate_fog(mdna_text)
        sheet.cell(row, 4, fog_score)
        row += 1

    profile.save('result.xlsx')


if __name__ == '__main__':
    main()
